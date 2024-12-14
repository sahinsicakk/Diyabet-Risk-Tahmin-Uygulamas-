from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import train_test_split

# CSV dosyasını oku
df = pd.read_csv('C:/Users/sahin/Desktop/Diabetes-Predict-main/diabetes.csv')

# KNN modelini eğitmek için veri hazırlığı
X = df.drop(columns='Outcome')
y = df['Outcome']
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

knn = KNeighborsClassifier(n_neighbors=3)
knn.fit(X_train, y_train)

# Referans değerleri
reference_ranges = {
    'Pregnancies': (0, 5),  # Gebelik sayısı genellikle 0-5 arasıdır, daha fazlası risk faktörü olabilir.
    'Glucose': (70, 99),     # Açlık kan şekeri 70-99 mg/dL normal kabul edilir.
    'BloodPressure': (90, 120),  # Normal kan basıncı 90/60 mmHg ile 120/80 mmHg arasındadır.
    'SkinThickness': (10, 30),   # Deri kalınlığı ölçümü, genellikle 10-30 mm arasında olmalıdır.
    'Insulin': (2, 25),     # Açlık insülin seviyesi genellikle 2-25 µIU/mL arasında olmalıdır.
    'BMI': (18.5, 24.9),    # Normal BMI aralığı 18.5 - 24.9 kg/m²'dir.
    'DiabetesPedigreeFunction': (0, 1.5),  # Diyabet soygeçmişi fonksiyonu 0-1.5 arasında değişir.
    'Age': (18, 75)         # Diyabet riski genellikle 18 yaş ve üstü için değerlendirilir, 75 yaşa kadar yaygın.
}


# Tkinter arayüzü
def gonder():
    try:
        # Kullanıcıdan verileri alıyoruz
        name_surname = entry_name.get()  # Kullanıcı adı ve soyadı
        age = float(entry_age.get())
        pregnancies = float(entry_pregnancies.get())
        glucose = float(entry_glucose.get())
        bloodpressure = float(entry_bloodpressure.get())
        skinthickness = float(entry_skinthickness.get())
        insulin = float(entry_insulin.get())
        bmi = float(entry_bmi.get())
        dpf = float(entry_dpf.get())

        # Kullanıcıdan alınan verileri uygun formata sokuyoruz
        Xn = [[pregnancies, glucose, bloodpressure, skinthickness, insulin, bmi, dpf, age]]

        # Tahmin yapma
        out = knn.predict(Xn)
        result_outcome.set(f"Sonuç Değeri: {out[0]}")  # result_outcome burada kullanılacak
        if out == 0:
            result_outcome.set("Durum: Diyabet Riski Yok")  # Label'ı güncelleme
        else:
            result_outcome.set("Durum: Diyabet Riski Var")  # Label'ı güncelleme

        # Referans dışı değerlerin kontrolü
        data = {
            'Pregnancies': pregnancies,
            'Glucose': glucose,
            'BloodPressure': bloodpressure,
            'SkinThickness': skinthickness,
            'Insulin': insulin,
            'BMI': bmi,
            'DiabetesPedigreeFunction': dpf,
            'Age': age
        }

        out_of_range = {}
        recommendations = {
            'Pregnancies': "Gebelik sayısını dengelemek, sağlıklı bir yaşam tarzı sürdürmek...",
            'Glucose': "Glikoz seviyesini kontrol altında tutmak için şekerli yiyeceklerden kaçının...",
            'BloodPressure': "Düşük sodyumlu bir diyet uygulayın, düzenli egzersiz yapın...",
            'SkinThickness': "Cilt sağlığını koruyacak bir beslenme uygulamak...",
            'Insulin': "Yüksek insülin seviyelerini dengelemek için diyetinizi gözden geçirin...",
            'BMI': "Daha düşük BMI elde etmek için dengeli bir diyet uygulayın...",
            'DiabetesPedigreeFunction': "Ailenizde diyabet geçmişi varsa, yaşam tarzınızı değiştirmek...",
            'Age': "Yaşla birlikte diyabet riskinin arttığını göz önünde bulundurarak sağlıklı yaşam alışkanlıkları edinin..."
        }

        for column, (low, high) in reference_ranges.items():
            if data[column] < low or data[column] > high:
                out_of_range[column] = data[column]

        # Görselleştirme
        plt.figure(figsize=(12, 8))
        for column, value in out_of_range.items():
            plt.scatter([column], [value], color="red", label=f"{column}: {value}")
            plt.axhline(reference_ranges[column][0], color="blue", linestyle="--", linewidth=1)
            plt.axhline(reference_ranges[column][1], color="blue", linestyle="--", linewidth=1)
        plt.title("Referans Dışı Değerler")
        plt.legend()
        plt.show()

        # Excel dosyasına yazma
        df_out_of_range = pd.DataFrame([data])
        df_out_of_range = df_out_of_range[[col for col in data.keys() if col in out_of_range]]
        
        # Ekstra önerileri Excel dosyasına ekliyoruz
        recommendations_df = pd.DataFrame.from_dict(recommendations, orient='index', columns=["Öneri"])

        # Kullanıcı adı soyadı ile başlık ekleyerek yazma
        excel_filename = f"{name_surname}.xlsx"
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            recommendations_df.to_excel(writer, sheet_name='Öneriler', startrow=1)
            df_out_of_range.to_excel(writer, index=False, sheet_name="Referans Dışı Değerler")

            # Diyabetle ilgili bilgi eklemek
            workbook = writer.book
            worksheet = workbook["Referans Dışı Değerler"]
            worksheet.cell(len(df_out_of_range) + 2, 1, f"Sayın {name_surname},")
            worksheet.cell(len(df_out_of_range) + 4, 1, "Diyabetsiz bir yaşam mümkün!")

            # Sonuçlar sayfası oluşturulması
            result_page = workbook.create_sheet("Sonuçlar")
            result_page.cell(row=1, column=1, value="Diyabet Riski Sonucu")
            result_page.cell(row=2, column=1, value="Durum")
            result_page.cell(row=2, column=2, value="Diyabet Riski Var" if out == 1 else "Diyabet Riski Yok")

            # Referans Değerleri sayfası eklemek
            reference_page = workbook.create_sheet("Referans Değerleri")
            reference_page.cell(row=1, column=1, value="Parametre")
            reference_page.cell(row=1, column=2, value="Referans Aralığı")

            # Referans değerlerini ekleme
            for idx, (param, (low, high)) in enumerate(reference_ranges.items(), start=2):
                reference_page.cell(row=idx, column=1, value=param)
                reference_page.cell(row=idx, column=2, value=f"{low} - {high}")

        result_outcome.set(f"Sonuçlar {excel_filename} dosyasına kaydedildi!")

    except ValueError as e:
        result_outcome.set(f"Lütfen geçerli bir sayı girin! Hata: {str(e)}")


# Tkinter arayüzü
root = Tk()
root.title("Diyabet Tahmin ve Referans Dışı Kontrol Uygulaması")

# Arkaplan resmi ekleme
background_image = Image.open("C:/Users/sahin/Desktop/Diabetes-Predict-main/diabetes.jpg")  # Arka plan görseli
background_image = background_image.resize((800,800), Image.Resampling.LANCZOS)  # Düzeltme yapıldı
background = ImageTk.PhotoImage(background_image)

background_label = Label(root, image=background)
background_label.place(relwidth=1, relheight=1)

# Giriş alanları
Label(root, text="İsim Soyisim :", bg="white", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=10, sticky="w")
entry_name = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_name.grid(row=0, column=1, padx=10, pady=10)

Label(root, text="Gebelik Sayısı :", bg="white", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=10, sticky="w")
entry_pregnancies = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_pregnancies.grid(row=1, column=1, padx=10, pady=10)

Label(root, text="Glikoz:", bg="white", font=("Arial", 12)).grid(row=2, column=0, padx=10, pady=10, sticky="w")
entry_glucose = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_glucose.grid(row=2, column=1, padx=10, pady=10)

Label(root, text="Kan Basıncı:", bg="white", font=("Arial", 12)).grid(row=3, column=0, padx=10, pady=10, sticky="w")
entry_bloodpressure = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_bloodpressure.grid(row=3, column=1, padx=10, pady=10)

Label(root, text="Cilt Kalınlığı:", bg="white", font=("Arial", 12)).grid(row=4, column=0, padx=10, pady=10, sticky="w")
entry_skinthickness = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_skinthickness.grid(row=4, column=1, padx=10, pady=10)

Label(root, text="İnsülin:", bg="white", font=("Arial", 12)).grid(row=5, column=0, padx=10, pady=10, sticky="w")
entry_insulin = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_insulin.grid(row=5, column=1, padx=10, pady=10)

Label(root, text="BMI:", bg="white", font=("Arial", 12)).grid(row=6, column=0, padx=10, pady=10, sticky="w")
entry_bmi = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_bmi.grid(row=6, column=1, padx=10, pady=10)

Label(root, text="Diabetes Pedigree Fonksiyonu:", bg="white", font=("Arial", 12)).grid(row=7, column=0, padx=10, pady=10, sticky="w")
entry_dpf = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_dpf.grid(row=7, column=1, padx=10, pady=10)

Label(root, text="Yaş:", bg="white", font=("Arial", 12)).grid(row=8, column=0, padx=10, pady=10, sticky="w")
entry_age = Entry(root, font=("Arial", 12), relief="solid", bd=2, bg="white", highlightthickness=0)
entry_age.grid(row=8, column=1, padx=10, pady=10)

# Sonuçlar için
result_outcome = StringVar()
Label(root, textvariable=result_outcome, bg="white", font=("Arial", 12, "bold")).grid(row=9, column=1, padx=10, pady=10)

# Gönder butonu
Button(root, text="Tahmin Et ve Kaydet", command=gonder, bg="#4CAF50", fg="white", font=("Arial", 12)).grid(row=10, column=1, padx=10, pady=20)

root.mainloop()
